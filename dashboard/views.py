from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.models import User, auth

from math import ceil
from openpyxl import load_workbook
from app_1.models import Stud   
import pandas as pd
from openpyxl.utils import get_column_letter
import datetime as dt
from app_1.models import Record
# Create your views here.
# wb = load_workbook('attendance.xlsx')
# sheet = wb.active
wb = None
sheet = None
 
record=[]
tot_lect=[]
tot_stud_on_a_day = []
def dashboard(request):
    if not request.user.is_authenticated:
        return redirect('login')
    wb = load_workbook('attendance.xlsx')
    sheet = wb.active
    record.clear()
    tot_lect.clear()
    tot_stud_on_a_day.clear()
    #tot lects taken by teacher
    # col=0
    # for c in sheet.iter_cols(min_row=1,max_row=1,min_col=3,max_col=sheet.max_column):
    #     col+=1
    # print(col)
    # print(sheet.max_column - 2)
    total_lecture = sheet.max_column - 2 # reduce 2 to remove the 1st two column count
    #retrieving data from db , calc tot_lect , display on html
    i=2
    for u in Stud.objects.all():
        data = {}
        #info of each student as a key of dict
        data['sid'] = u.id
        data['sname'] = u.name
        #tot lect of each student as a key of dict
        count=0
        for row in sheet[i]:
            if row.value == "P":
                count+=1
        data['stot_lect'] = count 
        #percentage claculation
        data['sper'] = ceil((count/total_lecture)*100)       
        #appending the count to list.
        tot_lect.append(count)
        i+=1
        record.append(data)
    
    for col in sheet.iter_cols(min_row=2,
                           max_row=sheet.max_row,
                           min_col=3,
                           max_col=sheet.max_column):
        cou = 0
        for data in col:
            if data.value == 'P':
                # print(data.value,end="")
                cou+=1        
        tot_stud_on_a_day.append(cou)
    print(tot_stud_on_a_day)
    
    # record.append({'sid':1, 'sname':'sarthak','stot_lect': 2,'sper':'78'})
    
    total_students = sheet.max_row - 1
    today_date = dt.datetime.now()
    today_date = f'{today_date:%d-%m-%Y}'
    course_code = Record.objects.filter(email = request.user.email).first()
    course_code = course_code.course_code
    
    # print(request.user.email)
    return render(request,'dashboard/index.html',
    {
    'record':record,
    'total_lecture':total_lecture,
    'total_students':total_students,
    'course_code':course_code,
    'today_date':today_date,
    'graph_data':tot_stud_on_a_day
    })
    
def download(request):
    #download the excel file and append the tot lect col
    wb = load_workbook('attendance.xlsx')
    sheet = wb.active
    new_column = get_column_letter(sheet.max_column + 1)
    header_cell = new_column + '1'
    sheet[header_cell] = 'Total Lecture'   
    print(tot_lect)
    for i, value in enumerate(tot_lect, start=2):
        cell = new_column + str(i)
        sheet[cell] = value
        print(value)  
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Student_Data.xlsx'
    wb.save(response)
    wb.save('Student_Data.xlsx')
    return response
   
   