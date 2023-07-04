from django.shortcuts import render
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.models import User, auth
from app_1.models import Record, Stud
# import openpyxl
# student_id = [1, 2, 3, 4, 7,9,10] # <- replace it here with the data received by the server
# date = "14-05-2023"
# workbook = openpyxl.load_workbook('attendance.xlsx')
# sheet = workbook.active
# val = sheet.max_column+1
# sheet.cell(1, val).value = date
# for i in student_id:
#     for row in range(2, sheet.max_row + 1):
#         key = int(sheet.cell(row, 1).value)
#         print(i)
#         if key == i:
#             sheet.cell(row, val).value = "Present"
# workbook.save('attendance.xlsx')

def mark(request): #http://127.0.0.1:8000/mark?rfid=1234  #rf_10 -teacher ,  rf_1 , rf_2, rf_3 for student
    if request.method == 'GET':
        try:
            rfid_number = request.GET['rfid']
            #can also use request.GET.get() functioni it returns None if parameter not present
            #also no need to put that in try block
        except Exception as e:
            rfid_number = "-1"
        # print(data)
        try:
            student = Stud.objects.get(rfid=rfid_number) 
        except Exception as e:
            student = None
        if student is not None:
            print(student) # mark attendance
            save_roll(str(student.roll))
        else:
            try:
                teacher = Record.objects.get(rfid=rfid_number)
            except Exception as e:
                teacher = None
            if teacher is not None:
                print(teacher) # save the student attendance    
                roll_list = load_file()
                print(roll_list)
                save_excel(roll_list) # save the student data into excel sheet
            else:
                return HttpResponse("0")
                print("Not found")

    return HttpResponse("1")
    return HttpResponse("Attendance marked" + rfid_number + 
    (student.name if student else (teacher.name if teacher else "No one") ))

def save_roll(roll: str):
    roll_list:list = load_file()    
    if roll in roll_list:
        roll_list.remove(roll)
    else:
        roll_list.append(roll)
    with open('temp.txt','w') as my_file:        
        my_file.write(','.join(roll_list))

def load_file():
    roll_list: list=[]
    try:   
        with open('temp.txt','r') as my_file:
            data = my_file.readline()
            roll_list = data.split(',')
            if('') in roll_list:
                roll_list.remove('')
    except Exception as e:
        with open('temp.txt','w') as temp:
            pass
    return roll_list

import openpyxl
import datetime as dt
def save_excel(student_id):
    
    # student_id = [1, 2, 3, 4, 7,9,10] # <- replace it here with the data received by the server
    date = dt.datetime.now()
    date = f"{date:%d-%m-%Y}"
    workbook = openpyxl.load_workbook('attendance.xlsx')
    sheet = workbook.active
    val = sheet.max_column+1
    sheet.cell(1, val).value = date           
    for row in range(2, sheet.max_row + 1):
        key = int(sheet.cell(row, 1).value)
        # print(i)
        if str(key) in student_id:
            print(key)
            sheet.cell(row, val).value = "P"
        else:
            sheet.cell(row, val).value = "A"
            
    workbook.save('attendance.xlsx')

def test(request):
    return HttpResponse("Hehe")